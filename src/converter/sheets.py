from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import TYPE_CHECKING

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font

from converter.directions import DIRECTION_COLUMN, normalize_direction
from converter.housing_types import HOUSING_TYPE_COLUMN, normalize_housing_type
from converter.prices import PRICE_COLUMNS, normalize_price

if TYPE_CHECKING:
    from google.oauth2.credentials import Credentials

    from converter.schema import ProjectType


_HYPERLINK_SEP = " → "

# Targets that get a literal fallback value when admin doesn't map a source column.
_DEFAULT_UNMAPPED_VALUES = {
    "Trạng thái căn hộ": "Công khai",
}


def _split_value_url(cell: str) -> tuple[str, str | None]:
    """Cell may be formatted as 'value → url' (set in _extract_visible_rows)."""
    if _HYPERLINK_SEP in cell:
        value, url = cell.split(_HYPERLINK_SEP, 1)
        return value.strip(), url.strip() or None
    return cell, None


def _norm_col_name(s: str) -> str:
    """Collapse all whitespace (incl. CRLF from HTML form encoding) to single space.

    HTML5 form submission converts in-value '\\n' to '\\r\\n', so a header like
    'Giá TTS\\n(tạm tính)' read from Sheets won't match the mapping value coming
    back from the browser. Normalising both sides on lookup avoids the mismatch.
    """
    return " ".join(s.split())


def build_xlsx(
    source_header: list[str],
    source_rows: list[list[str]],
    target_columns: tuple[str, ...],
    mapping: dict[str, str],
    literals: dict[str, str] | None = None,
    project_type: ProjectType | None = None,
) -> bytes:
    """Generate xlsx where columns follow target_columns.

    Resolution order per cell:
      1. `literals[target]` — same constant value across all rows;
      2. `source_row[mapping[target]]` — lookup from source data;
      3. `_DEFAULT_UNMAPPED_VALUES[target]` — static fallback.
    """
    literals = literals or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    bold = Font(bold=True)
    for col_idx, target in enumerate(target_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=target)
        cell.font = bold

    header_index = {_norm_col_name(col): i for i, col in enumerate(source_header)}

    for row_num, row_data in enumerate(source_rows, start=2):
        for col_idx, target in enumerate(target_columns, start=1):
            literal = literals.get(target)
            if literal:
                value = literal
                if target in PRICE_COLUMNS:
                    value = normalize_price(value)
                elif target == HOUSING_TYPE_COLUMN and project_type is not None:
                    value = normalize_housing_type(value, project_type)
                elif target == DIRECTION_COLUMN:
                    value = normalize_direction(value)
                ws.cell(row=row_num, column=col_idx, value=value)
                continue

            src_col = _norm_col_name(mapping.get(target) or "")
            if not src_col:
                default = _DEFAULT_UNMAPPED_VALUES.get(target)
                if default is not None:
                    ws.cell(row=row_num, column=col_idx, value=default)
                continue
            h_idx = header_index.get(src_col)
            if h_idx is None or h_idx >= len(row_data):
                continue
            raw = row_data[h_idx]
            value, url = _split_value_url(raw)
            # For "Link *" target columns: always display the URL as text. Source
            # cells like "TĐ11-56 → https://..." should write just the URL.
            # For other targets: only fall back to URL when there's no display text.
            if url and (target.startswith("Link ") or not value):
                value = url
            if target in PRICE_COLUMNS:
                value = normalize_price(value)
            elif target == HOUSING_TYPE_COLUMN and project_type is not None:
                value = normalize_housing_type(value, project_type)
            elif target == DIRECTION_COLUMN:
                value = normalize_direction(value)
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")
_GID_RE = re.compile(r"[?#&]gid=(\d+)")


class SheetURLError(ValueError):
    pass


class SheetReadError(RuntimeError):
    pass


@dataclass(frozen=True)
class SheetRef:
    sheet_id: str
    gid: int


def parse_url(url: str) -> SheetRef:
    if "docs.google.com" not in url:
        raise SheetURLError("URL phải là docs.google.com")
    id_match = _SHEET_ID_RE.search(url)
    if not id_match:
        raise SheetURLError("Không tìm thấy spreadsheet ID trong URL")
    gid_match = _GID_RE.search(url)
    return SheetRef(
        sheet_id=id_match.group(1),
        gid=int(gid_match.group(1)) if gid_match else 0,
    )


def detect_header(rows: list[list[str]]) -> list[str]:
    for row in rows:
        non_empty = [cell.strip() for cell in row if cell.strip()]
        if len(non_empty) < 4:
            continue
        text_cells = sum(1 for cell in non_empty if not cell[0].isdigit())
        if text_cells >= len(non_empty) * 0.7:
            return [cell.strip() for cell in row]
    for row in rows:
        if any(cell.strip() for cell in row):
            return [cell.strip() for cell in row]
    return []


_JUNK_CELL_MAX_LEN = 50


_CODE_PATTERN_RE = re.compile(r"^[A-ZĐ][A-ZĐa-zđ\d]*-\d", re.UNICODE)


def _is_header_candidate(row: list[str]) -> bool:
    """Row looks like a header: ≥4 non-empty cells, ≥70% start with non-digit,
    and no cell matches an ID code pattern like 'SX21-01' or 'VT64-07'."""
    non_empty = [cell.strip() for cell in row if cell.strip()]
    if len(non_empty) < 4:
        return False
    # Header cells never contain unit ID codes — these only appear in data rows.
    if any(_CODE_PATTERN_RE.match(c) for c in non_empty):
        return False
    text_cells = sum(1 for cell in non_empty if not cell[0].isdigit())
    return text_cells >= len(non_empty) * 0.7


def _is_data_row(row: list[str]) -> bool:
    """Data row must have ≥2 non-empty cells AND contain a unit ID code (Mã căn)."""
    cells = [c.strip() for c in row if c.strip()]
    if len(cells) < 2:
        return False
    return any(_CODE_PATTERN_RE.match(c) for c in cells)


def _merge_header_rows(main: list[str], sub: list[str]) -> list[str]:
    """Combine sub-header values into main using nearest non-empty main cell as parent."""
    merged: list[str] = []
    parent = ""
    width = max(len(main), len(sub))
    for i in range(width):
        m = (main[i] if i < len(main) else "").strip()
        s = (sub[i] if i < len(sub) else "").strip()
        if m:
            parent = m
        if s:
            merged.append(f"{parent} - {s}" if parent and parent != s else s)
        else:
            merged.append(m)
    return merged


def split_into_blocks(
    rows: list[list[str]],
) -> list[tuple[list[str], list[list[str]]]]:
    """Detect all (header, data_rows) blocks in a sheet.

    A new block starts at each header-like row. Sub-header rows (sparse + fill
    gaps) are merged into the parent header. Data rows pass the narrow junk
    filter (≥2 cells, no cell longer than _JUNK_CELL_MAX_LEN chars).
    """
    blocks: list[tuple[list[str], list[list[str]]]] = []
    cur_header: list[str] | None = None
    cur_data: list[list[str]] = []

    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        if _is_header_candidate(row):
            if cur_header is not None and cur_data:
                blocks.append((cur_header, cur_data))

            main = row
            next_i = i + 1
            if next_i < n:
                cand = rows[next_i]
                cand_non_empty = sum(1 for c in cand if c.strip())
                main_size = sum(1 for c in main if c.strip())
                fills_gaps = any(
                    j < len(main) and cand[j].strip() and not main[j].strip()
                    for j in range(len(cand))
                )
                if 0 < cand_non_empty < main_size * 0.5 and fills_gaps:
                    main = _merge_header_rows(main, cand)
                    next_i += 1

            # Keep header positionally aligned with data rows — don't drop empty
            # cells (e.g. leading blank columns A/B in some sheets), otherwise
            # header_index → row_data lookup gets off-by-N.
            cur_header = [c.strip() for c in main]
            cur_data = []
            i = next_i
            continue

        if cur_header is not None and _is_data_row(row):
            cur_data.append(row)
        i += 1

    if cur_header is not None and cur_data:
        blocks.append((cur_header, cur_data))

    return blocks


_GRID_FIELDS = (
    "sheets.data("
    "rowData.values("
    "formattedValue,effectiveFormat.backgroundColor,hyperlink,"
    "userEnteredValue,textFormatRuns,chipRuns"
    "),"
    "rowMetadata(hiddenByUser,hiddenByFilter),"
    "columnMetadata.hiddenByUser"
    ")"
)

# Rows tinted with this background mark sold/locked units and must be excluded.
_EXCLUDED_BG_HEX = "#ff0000"

_HYPERLINK_FORMULA_RE = re.compile(r'HYPERLINK\s*\(\s*"([^"]+)"', re.IGNORECASE)
_DEBUG_URL_FOUND = 0


def _resolve_cell_url(cell: dict) -> str | None:
    """Try every place Sheets stores a URL: cell-level link, formula, rich text, smart chip."""
    url = cell.get("hyperlink")
    if url:
        return url
    formula = cell.get("userEnteredValue", {}).get("formulaValue", "") or ""
    if formula:
        match = _HYPERLINK_FORMULA_RE.search(formula)
        if match:
            return match.group(1)
    for run in cell.get("textFormatRuns", []) or []:
        link_uri = run.get("format", {}).get("link", {}).get("uri")
        if link_uri:
            return link_uri
    for chip_run in cell.get("chipRuns", []) or []:
        chip = chip_run.get("chip", {})
        # Drive file chips
        uri = chip.get("richLinkProperties", {}).get("uri")
        if uri:
            return uri
        # Plain link chips (some versions)
        uri = chip.get("linkProperties", {}).get("uri")
        if uri:
            return uri
    return None


def _bg_hex(cell: dict) -> str | None:
    color = cell.get("effectiveFormat", {}).get("backgroundColor")
    if not color:
        return None
    r = int(round(color.get("red", 0) * 255))
    g = int(round(color.get("green", 0) * 255))
    b = int(round(color.get("blue", 0) * 255))
    return f"#{r:02x}{g:02x}{b:02x}"


def read_source_columns(
    url: str,
    credentials: Credentials,
    *,
    sample_rows: int = 30,
) -> list[str]:
    ref = parse_url(url)
    try:
        service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
        sheet_name = _resolve_sheet_name(service, ref.sheet_id, ref.gid)
        grid = (
            service.spreadsheets()
            .get(
                spreadsheetId=ref.sheet_id,
                ranges=[f"'{sheet_name}'!A1:Z{sample_rows}"],
                includeGridData=True,
                fields=_GRID_FIELDS,
            )
            .execute()
        )
    except HttpError as exc:
        raise SheetReadError(_format_api_error(exc)) from exc

    visible = _extract_visible_rows(grid)
    return [cell for cell in detect_header(visible) if cell]


def read_all_rows(url: str, credentials: Credentials) -> list[list[str]]:
    ref = parse_url(url)
    try:
        service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
        sheet_name = _resolve_sheet_name(service, ref.sheet_id, ref.gid)
        grid = (
            service.spreadsheets()
            .get(
                spreadsheetId=ref.sheet_id,
                ranges=[f"'{sheet_name}'"],
                includeGridData=True,
                fields=_GRID_FIELDS,
            )
            .execute()
        )
    except HttpError as exc:
        raise SheetReadError(_format_api_error(exc)) from exc

    return _extract_visible_rows(grid)


def _resolve_sheet_name(service, sheet_id: str, gid: int) -> str:
    metadata = (
        service.spreadsheets()
        .get(
            spreadsheetId=sheet_id,
            fields="sheets.properties(sheetId,title)",
        )
        .execute()
    )
    for sheet in metadata.get("sheets", []):
        props = sheet["properties"]
        if props["sheetId"] == gid:
            return props["title"]
    raise SheetReadError(f"Không tìm thấy tab có gid={gid} trong spreadsheet")


def _extract_visible_rows(api_result: dict) -> list[list[str]]:
    sheets = api_result.get("sheets", [])
    if not sheets:
        return []
    blocks = sheets[0].get("data", [])
    if not blocks:
        return []
    grid = blocks[0]
    row_meta = grid.get("rowMetadata", [])
    col_meta = grid.get("columnMetadata", [])
    hidden_cols = {i for i, m in enumerate(col_meta) if m.get("hiddenByUser")}
    rows = grid.get("rowData", [])
    visible: list[list[str]] = []
    for idx, row in enumerate(rows):
        meta = row_meta[idx] if idx < len(row_meta) else {}
        preview = [(cell.get("formattedValue") or "").strip() for cell in row.get("values", [])]
        row_has_code = any(_CODE_PATTERN_RE.match(c) for c in preview if c)
        # Skip hidden rows only when they look like data (have Mã căn code).
        # Header rows can also be marked hidden — keep them so blocks form.
        if (meta.get("hiddenByUser") or meta.get("hiddenByFilter")) and row_has_code:
            continue
        visible_cells = [
            cell for i, cell in enumerate(row.get("values", [])) if i not in hidden_cols
        ]
        # Skip red-bg rows only when they're data (sold/locked unit). Header rows
        # may have an individual red cell (e.g. styled "ĐÃ BÁN" column header)
        # without actually being a sold-unit row — keep them.
        if row_has_code and any(_bg_hex(cell) == _EXCLUDED_BG_HEX for cell in visible_cells):
            continue
        cells = []
        for cell in visible_cells:
            value = cell.get("formattedValue") or ""
            url = _resolve_cell_url(cell)
            # Keep BOTH value and url so the unit-ID regex can still match the
            # text portion (e.g. "BM3-26") even when the cell carries a hyperlink.
            cells.append(f"{value}{_HYPERLINK_SEP}{url}" if url else value)
        # Trim trailing format-only cells (banding/borders) that have no value
        while cells and not cells[-1].strip():
            cells.pop()
        if not cells:
            continue
        visible.append(cells)
    max_len = max((len(r) for r in visible), default=0)
    return [r + [""] * (max_len - len(r)) for r in visible]


def _format_api_error(exc: HttpError) -> str:
    status = exc.resp.status if exc.resp else "?"
    if status == 403:
        return (
            "Bạn không có quyền đọc sheet này. Kiểm tra account đã đăng nhập có được share không."
        )
    if status == 404:
        return "Sheet không tồn tại hoặc URL sai."
    if status == 401:
        return "Token Google hết hạn. Đăng xuất rồi đăng nhập lại."
    return f"Sheets API lỗi (HTTP {status}): {exc.reason or 'unknown'}"
