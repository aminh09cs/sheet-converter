from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from converter.schema import HIGH_RISE_COLUMNS, LOW_RISE_COLUMNS, ProjectType

_PROJECT_CODE_COLUMN = "Mã dự án"
_DUP_FILL = PatternFill(start_color="FFFFE082", end_color="FFFFE082", fill_type="solid")


class MergeError(ValueError):
    pass


def _detect_type(header: tuple[str, ...]) -> ProjectType:
    if header == HIGH_RISE_COLUMNS:
        return ProjectType.HIGH_RISE
    if header == LOW_RISE_COLUMNS:
        return ProjectType.LOW_RISE
    raise MergeError(
        "Header file không khớp schema chuẩn — chỉ chấp nhận file output từ tab Chuẩn hóa"
    )


def _read_xlsx(content: bytes) -> tuple[tuple[str, ...], list[list[tuple[Any, str | None]]]]:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise MergeError(f"Không đọc được file xlsx: {exc}") from exc
    ws = wb.active
    rows = list(ws.iter_rows())
    if not rows:
        raise MergeError("File trống")

    header_cells = rows[0]
    header_values: list[str] = []
    for cell in header_cells:
        v = cell.value
        header_values.append(str(v).strip() if v is not None else "")
    while header_values and not header_values[-1]:
        header_values.pop()
    header = tuple(header_values)

    data: list[list[tuple[Any, str | None]]] = []
    width = len(header)
    for row in rows[1:]:
        row_cells = list(row[:width])
        if not any(c.value is not None and str(c.value).strip() for c in row_cells):
            continue
        row_data: list[tuple[Any, str | None]] = []
        for cell in row_cells:
            url = cell.hyperlink.target if cell.hyperlink else None
            row_data.append((cell.value, url))
        while len(row_data) < width:
            row_data.append((None, None))
        data.append(row_data)

    return header, data


def merge_files(
    files: list[tuple[str, bytes]],
) -> tuple[ProjectType, tuple[str, ...], list[list[tuple[Any, str | None]]], set[str]]:
    """Validate uniform schema, concat rows, mark dup `Mã dự án`.

    Returns (type, header, rows, dup_codes).
    """
    if len(files) < 2:
        raise MergeError("Cần tối thiểu 2 file để merge")

    expected_type: ProjectType | None = None
    expected_header: tuple[str, ...] | None = None
    all_rows: list[list[tuple[Any, str | None]]] = []
    code_to_files: dict[str, set[str]] = defaultdict(set)

    for filename, content in files:
        header, data = _read_xlsx(content)
        file_type = _detect_type(header)

        if expected_type is None:
            expected_type = file_type
            expected_header = header
        elif file_type != expected_type:
            raise MergeError(
                f'"{filename}" là {file_type.label}, không khớp với loại của file đầu tiên ({expected_type.label})'
            )

        code_idx = header.index(_PROJECT_CODE_COLUMN)
        for row in data:
            value = row[code_idx][0] if code_idx < len(row) else None
            if value is not None and str(value).strip():
                code_to_files[str(value).strip()].add(filename)
            all_rows.append(row)

    assert expected_type is not None and expected_header is not None
    dup_codes = {code for code, fs in code_to_files.items() if len(fs) > 1}
    return expected_type, expected_header, all_rows, dup_codes


def build_merged_xlsx(
    header: tuple[str, ...],
    rows: list[list[tuple[Any, str | None]]],
    dup_codes: set[str],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"

    bold = Font(bold=True)
    for col_idx, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold

    code_idx = header.index(_PROJECT_CODE_COLUMN)

    for row_num, row_data in enumerate(rows, start=2):
        code_value = row_data[code_idx][0] if code_idx < len(row_data) else None
        is_dup = code_value is not None and str(code_value).strip() in dup_codes
        for col_idx, (value, url) in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
            if is_dup:
                cell.fill = _DUP_FILL

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
