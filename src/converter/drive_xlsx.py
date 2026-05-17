"""Fallback path for reading xlsx files uploaded to Drive.

The Sheets API only works on NATIVE Google Sheets. When admin pastes a link
to an uploaded .xlsx file, the API returns HTTP 400 with
"This operation is not supported for this document. The document must not be
an Office file." This module downloads the file via Drive API and parses it
with openpyxl.

Note: openpyxl-parsed rows lose some metadata vs the Sheets API path
(rows hidden by basic filter are NOT detected, smart-chip URLs may not parse).
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

if TYPE_CHECKING:
    from google.oauth2.credentials import Credentials
    from googleapiclient.errors import HttpError


_HYPERLINK_SEP = " → "
# Solid red marks sold/locked units — same convention as the Sheets API path.
_EXCLUDED_BG_HEX = "ff0000"


def is_office_file_error(exc: HttpError) -> bool:
    """True if the Sheets API rejection was because the file is an uploaded xlsx."""
    if exc.resp is None or exc.resp.status != 400:
        return False
    text = str(exc).lower()
    return "office file" in text or "not supported for this document" in text


def read_xlsx_via_drive(file_id: str, credentials: Credentials) -> list[list[str]]:
    """Download the xlsx file from Drive and extract visible rows."""
    service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    request = service.files().get_media(fileId=file_id)
    buffer = BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return _parse_xlsx(buffer.getvalue())


def _parse_xlsx(content: bytes) -> list[list[str]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    hidden_rows = {i for i, dim in ws.row_dimensions.items() if dim.hidden}
    hidden_cols: set[int] = set()
    for letter, dim in ws.column_dimensions.items():
        if dim.hidden:
            hidden_cols.add(column_index_from_string(letter) - 1)

    visible: list[list[str]] = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx in hidden_rows:
            continue
        if any(_is_excluded_bg(cell) for cell in row):
            continue
        cells: list[str] = []
        for col_idx, cell in enumerate(row):
            if col_idx in hidden_cols:
                continue
            value = "" if cell.value is None else str(cell.value)
            url = cell.hyperlink.target if cell.hyperlink else None
            # Keep both value and url so unit-ID regex can still match text portion.
            cells.append(f"{value}{_HYPERLINK_SEP}{url}" if url else value)
        while cells and not cells[-1].strip():
            cells.pop()
        if not cells:
            continue
        visible.append(cells)
    max_len = max((len(r) for r in visible), default=0)
    return [r + [""] * (max_len - len(r)) for r in visible]


def _is_excluded_bg(cell) -> bool:
    """True if cell has solid red background (sold/locked marker)."""
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type != "solid":
        return False
    color = fill.start_color
    if not color:
        return False
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return False
    if len(rgb) == 8:
        rgb = rgb[2:]
    return rgb.lower() == _EXCLUDED_BG_HEX
